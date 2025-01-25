import os
import time
import sys
import requests
import pandas as pd
from dotenv import load_dotenv
from loguru import logger
import ipaddress

# For coloring cells / applying formats
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------------------------------------------------------------------------
# Configure the logger (Loguru)
# ------------------------------------------------------------------------------------
logger.remove()  # Remove Loguru's default handler
logger.add(
    sys.stderr,
    format=(
        "<green>{time:YYYY-MM-DD HH:mm:ss}</green> | "
        "<level>{level: <8}</level> | "
        "<cyan>{name}</cyan>:<cyan>{function}</cyan>:<cyan>{line}</cyan> - "
        "<level>{message}</level>"
    ),
    colorize=True,
    level="INFO"
)

# ------------------------------------------------------------------------------------
# Load environment variables from .env
# ------------------------------------------------------------------------------------
load_dotenv()

API_KEY = os.getenv("API_KEY")
BASE_URL = os.getenv("BASE_URL", "https://api.thousandeyes.com/v7")

# Common headers for most calls
HEADERS = {
    "Authorization": f"Bearer {API_KEY}",
    "Content-Type": "application/json"
}

# Special headers for the Endpoint Agents call (accepting HAL+JSON)
HEADERS_HAL = HEADERS.copy()
HEADERS_HAL["Accept"] = "application/hal+json"


def _apply_color_fills(excel_file: str, sheet_name: str = "Labels",
                       color_col: str = "color"):
    """
    Open the given Excel file with openpyxl and apply cell background
    colors in the given sheet, based on the hex codes in `color_col`.

    - If color_col has #93249F or 93249F, we fill the cell with that color.
    - Hex must be exactly 6 characters after removing the optional '#' or we skip.
    """
    logger.info(
        f"Applying cell color fills in '{sheet_name}' using column '{color_col}'...")
    wb = load_workbook(excel_file)

    if sheet_name not in wb.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found; skipping color fills.")
        wb.close()
        return

    ws = wb[sheet_name]

    # Find the column index for color_col.
    col_index = None
    for cell in ws[1]:  # row 1 => headers
        if cell.value == color_col:
            col_index = cell.column  # 1-based index
            break

    if col_index is None:
        logger.warning(
            f"Column '{color_col}' not found in '{sheet_name}' sheet; skipping color fills.")
        wb.close()
        return

    # Iterate from row 2 downward because row 1 is the header
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_index)
        color_val = str(cell.value).strip() if cell.value else ""
        if not color_val:
            continue

        # Remove a leading "#" if present
        if color_val.startswith("#"):
            color_val = color_val[1:]

        # Must be exactly 6 hex digits (case-insensitive)
        color_val = color_val.upper()
        if len(color_val) == 6:
            fill = PatternFill(start_color=color_val, end_color=color_val,
                               fill_type="solid")
            cell.fill = fill
        else:
            logger.debug(f"Skipping invalid color code '{cell.value}' in row {row_idx}.")

    wb.save(excel_file)
    wb.close()
    logger.info(f"Cell coloring applied; file re-saved as '{excel_file}'.")


def _auto_format_numbers(excel_file: str):
    """
    After writing the DataFrame to Excel, open it with openpyxl and
    detect columns that are entirely numeric. For those columns,
    set 'number_format' to 'General' (or '0', '#,##0.00', etc.).

    This ensures Excel interprets them as numbers rather than text.

    Rules:
    - We skip the header row (row 1).
    - If every non-empty cell in a column can be cast to float, we mark that column numeric.
    - Then we apply 'number_format' to each cell in that column from row 2 downward.
    """
    logger.info("Auto-formatting numeric columns in all sheets...")
    wb = load_workbook(excel_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.info(f"  Checking numeric columns in sheet '{sheet_name}'...")

        # We'll find how many columns we have by checking row 1
        max_col = ws.max_column
        # row 1 is the header
        for col_idx in range(1, max_col + 1):
            is_numeric = True
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None or val == "":
                    # skip empty
                    continue
                # If it's already a number in Python, that's good
                if isinstance(val, (int, float)):
                    continue
                # If it's a string, attempt to parse float
                if isinstance(val, str):
                    try:
                        float(val.replace(",", ""))  # remove comma if present
                    except ValueError:
                        is_numeric = False
                        break
                else:
                    # Not int, not float, not string => not numeric
                    is_numeric = False
                    break

            # If still numeric, apply a number format
            if is_numeric:
                # We'll do it from row 2 to last row
                for row_idx in range(2, ws.max_row + 1):
                    num_cell = ws.cell(row=row_idx, column=col_idx)
                    # Set a number format
                    # Could use '0', '#,##0', '#,##0.00', etc.
                    num_cell.number_format = "General"

    wb.save(excel_file)
    wb.close()
    logger.info(f"Numeric columns formatted in '{excel_file}'.")


def ip_in_any_subnet(ip_str, subnet_str_list):
    """
    Return True if ip_str (e.g. "10.0.0.130") is inside
    any of the subnets in subnet_str_list (e.g. ["10.228.0.0/17"]).
    """
    try:
        ip_obj = ipaddress.ip_address(ip_str)
    except ValueError:
        return False

    for net_str in subnet_str_list:
        try:
            net = ipaddress.ip_network(net_str, strict=False)
        except ValueError:
            # If it's an invalid network string, skip
            continue
        if ip_obj in net:
            return True

    return False


def agent_matches_filter(agent_row, filter_key, filter_values, mode="in"):
    """
    Checks if `agent_row` satisfies a single filter dict, e.g.:
      {
        "key": "username",
        "values": ["HCA\\xyz", "HCA\\abc"],
        "mode": "in"
      }
    Return True if it passes, False otherwise.

    `agent_row` is a Series from endpoint_agents_df, containing:
      - agentRow["id"]
      - agentRow["usernames"]
      - agentRow["localIpv4"]
      - agentRow["hardwareTypes"]
      - ... etc.

    For 'in' mode, we just check if there's an intersection or exact match.
    """
    # We'll handle a few known filter keys:
    filter_key = filter_key.lower().strip()
    if mode.lower() != "in":
        # For simplicity, only demonstrate 'in'.
        # If you have 'not_in' or others, handle them similarly.
        return False

    # 1) agent-id
    if filter_key == "agent-id":
        # agentRow["id"] is a single string
        agent_id = str(agent_row["id"]).strip()
        return agent_id in filter_values

    # 2) username
    elif filter_key == "username":
        # agentRow["usernames"] might be "HCA\\abc,HCA\\xyz"
        # split them into a set:
        agent_usernames = {u.strip() for u in agent_row["usernames"].split(",") if u}
        # check if any overlap
        return bool(agent_usernames.intersection(filter_values))

    # 3) local-network
    elif filter_key == "local-network":
        # filter_values might be subnets like ["10.228.0.0/17", "192.168.1.0/24"]
        # agentRow["localIpv4"] might be "10.228.5.100,192.168.1.55"
        ip_list = [ip.strip() for ip in agent_row["localIpv4"].split(",") if ip]
        for ip_str in ip_list:
            if ip_in_any_subnet(ip_str, filter_values):
                return True
        return False

    # 4) connection
    elif filter_key == "connection":
        # agentRow["hardwareTypes"] might be "ethernet,wireless"
        agent_hw = {t.strip().lower() for t in agent_row["hardwareTypes"].split(",") if t}
        # filter_values might be ["Ethernet"] or ["Ethernet", "Wireless"]
        filter_hw = {v.lower() for v in filter_values}
        # check if there's an intersection
        return bool(agent_hw.intersection(filter_hw))

    # 5) vpn-vendor or vpnType
    # If you store "vpnType" in agentRow, or parse agentRow["vpnInfo"],
    # you'll need to check that. For example:
    elif filter_key == "vpn-vendor" or filter_key == "vpnType":
        # If you have a dedicated column agentRow["vpnType"], easy:
        # return agentRow["vpnType"] in filter_values
        #
        # If you stored multiple in agentRow["vpnInfo"], you might parse them out
        # or check for strings like "vpnType=cisco-anyconnect" ...
        # For demonstration, just do a partial check:
        vpninfo_str = agent_row["vpnInfo"] or ""
        # e.g. "vpnType=cisco-anyconnect,gateway=1.2.3.4,client=10.1.2.3"
        # We'll see if any 'filter_val' is in the line
        for fv in filter_values:
            if fv.lower() in vpninfo_str.lower():
                return True
        return False

    # fallback
    return False


def agent_matches_label(agent_row, label_row):
    """
    Return True if the agent matches all filters (if matchType=and)
    or at least one filter (if matchType=or).
    """
    match_type = (label_row["matchType"] or "").lower().strip()
    if not match_type:
        match_type = "and"  # default if missing

    # The label row might have columns for each filter key. For example:
    #   label_row["agent_id_filter"] => "1234-abc...,5678-def..."
    #   label_row["username_filter"] => "HCA\\abc,HCA\\xyz"
    # We can parse each of these if non-empty into a single "filter" dict
    # that calls agent_matches_filter.

    # Build a list of filter dicts from the label row.
    # (We do this because in your script, you store each key as a comma string).
    filters_to_check = []

    # agent_id_filter
    if label_row.get("agent_id_filter"):
        # split by comma
        vals = [v.strip() for v in label_row["agent_id_filter"].split(",") if v.strip()]
        filters_to_check.append({
            "key": "agent-id",
            "values": vals,
            "mode": "in"
        })

    # username_filter
    if label_row.get("username_filter"):
        vals = [v.strip() for v in label_row["username_filter"].split(",") if v.strip()]
        filters_to_check.append({
            "key": "username",
            "values": vals,
            "mode": "in"
        })

    # local_network_filter
    if label_row.get("local_network_filter"):
        vals = [v.strip() for v in label_row["local_network_filter"].split(",") if v.strip()]
        filters_to_check.append({
            "key": "local-network",
            "values": vals,
            "mode": "in"
        })

    # vpn_vendor_filter
    if label_row.get("vpn_vendor_filter"):
        vals = [v.strip() for v in label_row["vpn_vendor_filter"].split(",") if v.strip()]
        filters_to_check.append({
            "key": "vpn-vendor",
            "values": vals,
            "mode": "in"
        })

    # connection_filter
    if label_row.get("connection_filter"):
        vals = [v.strip() for v in label_row["connection_filter"].split(",") if v.strip()]
        filters_to_check.append({
            "key": "connection",
            "values": vals,
            "mode": "in"
        })

    # Evaluate them
    if match_type == "and":
        # ALL must pass
        return all(agent_matches_filter(agent_row, f["key"], f["values"], f["mode"]) for f in filters_to_check)
    else:
        # OR => any pass
        return any(agent_matches_filter(agent_row, f["key"], f["values"], f["mode"]) for f in filters_to_check)


def build_label_agents_map(labels_df, endpoint_agents_df):
    """
    Returns a dict: { label_id: set_of_agent_ids }
    by checking each label's filters against each agent.
    """
    label_agents = {}

    for lbl_idx, lbl_row in labels_df.iterrows():
        label_id = lbl_row["id"]
        matching_agents = set()

        for ag_idx, ag_row in endpoint_agents_df.iterrows():
            agent_id = ag_row["id"]
            if agent_matches_label(ag_row, lbl_row):
                matching_agents.add(agent_id)

        label_agents[label_id] = matching_agents

    return label_agents


def build_test_agents_map(scheduled_tests_df, label_agents_map, endpoint_agents_df):
    """
    Returns a dict: { testID: set_of_agent_ids }
    """
    all_agent_ids = set(endpoint_agents_df["id"].unique())

    test_agents = {}
    for idx, test_row in scheduled_tests_df.iterrows():
        t_id = test_row["testID"]
        sel_type = test_row["agentSelectorType"]

        assigned_ids = set()
        if sel_type == "all-agents":
            assigned_ids = set(all_agent_ids)

        elif sel_type == "specific-agents":
            # parse test_row["assignedAgentsRaw"] -> "agent-id1,agent-id2"
            raw = test_row.get("assignedAgentsRaw", "")
            assigned_ids = {a.strip() for a in raw.split(",") if a.strip()}

        elif sel_type == "agent-labels":
            # parse test_row["assignedLabelsRaw"] -> "140737488492028,140737488493860"
            raw = test_row.get("assignedLabelsRaw", "")
            label_ids = {lbl.strip() for lbl in raw.split(",") if lbl.strip()}
            for lid in label_ids:
                # union with label_agents_map[lid]
                if lid in label_agents_map:
                    assigned_ids |= label_agents_map[lid]

        test_agents[t_id] = assigned_ids

    return test_agents


def get_account_ids(file_path: str) -> pd.DataFrame:
    """
    Reads an Excel file containing account IDs in one tab.
    Expected columns:
        - accountGroupName
        - aid
    Returns a DataFrame with columns: [accountGroupName, aid]
    """
    logger.info(f"Reading account IDs from '{file_path}'...")
    df = pd.read_excel(file_path, sheet_name=0)  # read the first sheet
    # Ensure required columns are present
    required_cols = {"accountGroupName", "aid"}
    if not required_cols.issubset(df.columns):
        logger.error("Excel file must contain 'accountGroupName' and 'aid' columns.")
        raise ValueError("Excel must contain 'accountGroupName' and 'aid' columns.")
    logger.success("Account IDs read successfully.")
    return df[["accountGroupName", "aid"]]


def build_test_agent_assignment_df(scheduled_tests_df, test_agents_map, endpoint_agents_df):
    """
    Produces a DataFrame with columns [testID, testName, agentID, agentName].
    """
    # Make a quick lookup for agentID -> agentName
    agent_lookup = endpoint_agents_df.set_index("id")["name"].to_dict()
    # Also a quick lookup for testID -> testName
    test_lookup = scheduled_tests_df.set_index("testID")["testName"].to_dict()

    rows = []
    for test_id, agent_ids in test_agents_map.items():
        tname = test_lookup.get(test_id, "")
        for ag_id in agent_ids:
            ag_name = agent_lookup.get(ag_id, "")
            rows.append({
                "testID": test_id,
                "testName": tname,
                "agentID": ag_id,
                "agentName": ag_name
            })

    return pd.DataFrame(rows)


def fetch_agents(aid: str) -> pd.DataFrame:
    """
    Fetch Enterprise Agents for a given Account ID.
    Endpoint: /agents?aid=XXXXXXX
    """
    url = f"{BASE_URL}/agents?aid={aid}"
    logger.info(f"Fetching Agents for AID={aid} from '{url}'...")
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()

    data = resp.json().get("agents", [])
    logger.info(f"Received {len(data)} Agents for AID={aid}.")
    records = []
    for agent in data:
        records.append({
            "OrgId": agent.get("orgId"),
            "agentId": agent.get("agentId"),
            "agentName": agent.get("agentName"),
            "agentType": agent.get("agentType"),
            "agentState": agent.get("agentState"),
            "lastSeen": agent.get("lastSeen"),
            "createdDate": agent.get("createdDate"),
            "utilization": agent.get("utilization"),
            "location": agent.get("location"),
            "enabled": agent.get("enabled"),
            "hostname": agent.get("hostname"),
            "ipAddresses": ", ".join(agent.get("ipAddresses", []))
            if isinstance(agent.get("ipAddresses"), list) else ""
        })
    return pd.DataFrame(records)


def fetch_endpoint_agents(aid: str) -> pd.DataFrame:
    """
    Fetch Endpoint Agents for a given Account ID, with expanded details
    from 'clients', 'vpnProfiles', and 'networkInterfaceProfiles'.

    Endpoint: /endpoint/agents?aid={aid}&expand=clients,vpnProfiles,networkInterfaceProfiles

    Returns columns:
    [
      "id", "name", "computerName", "osVersion", "platform", "lastSeen", "status",
      "deleted", "version", "createdAt", "numberOfClients", "locationName",
      "agentType", "licenseType",
      "usernames", "localIpv4", "gatewayIpv4", "hardwareTypes", "vpnInfo"
      ... (add more if needed)
    ]
    """
    # Construct URL with expansions
    next_url = (
        f"{BASE_URL}/endpoint/agents?aid={aid}"
        f"&expand=clients,vpnProfiles,networkInterfaceProfiles"
    )
    logger.info(
        f"Fetching Endpoint Agents for AID={aid} from '{next_url}' using HAL+JSON headers..."
    )

    all_records = []
    page_count = 1

    while next_url:
        resp = requests.get(next_url, headers=HEADERS_HAL)
        resp.raise_for_status()

        data = resp.json()
        agents_list = data.get("agents", [])

        links = data.get("_links", {})
        next_link = links.get("next", {}).get("href")

        logger.info(
            f"Page {page_count}: Received {len(agents_list)} Endpoint Agents. "
            f"{'Continuing...' if next_link else 'No more pages.'}"
        )

        for agent in agents_list:
            location_obj = agent.get("location", {}) or {}
            location_name = location_obj.get("locationName", "")

            # 1) Collect any usernames found in the 'clients' array
            username_set = set()
            for c in agent.get("clients", []):
                user_profile = c.get("userProfile", {})
                uname = user_profile.get("userName")
                if uname:
                    username_set.add(uname.strip())

            # 2) Parse networkInterfaceProfiles to gather local IPv4 addresses, gateways, hardware types
            local_ipv4_set = set()
            gateway_set = set()
            hardware_set = set()

            for nic in agent.get("networkInterfaceProfiles", []):
                hardware_type = nic.get("hardwareType", "")
                if hardware_type:
                    hardware_set.add(hardware_type)

                address_profiles = nic.get("addressProfiles", [])
                for ap in address_profiles:
                    # 'addressType' can be "unique-local", "unique-global", etc.
                    # We'll check if the ipAddress is a private IPv4 (e.g., 10.x, 192.168.x, 172.16-31.x)
                    ip_addr = ap.get("ipAddress", "")
                    gateway = ap.get("gateway", "")
                    if gateway:
                        gateway_set.add(gateway)

                    # Check if ip_addr is private IPv4
                    # A quick approach is to check if it starts with 10., 192.168., or 172.(16-31).
                    if is_private_ipv4(ip_addr):
                        local_ipv4_set.add(ip_addr)

            # 3) Collect VPN details if any
            vpn_info_list = []
            for vpnp in agent.get("vpnProfiles", []):
                # Example: "vpnType": "cisco-anyconnect"
                #          "vpnGatewayAddress": "165.214.12.240"
                #          "vpnClientAddresses": ["10.155.159.173"]
                #          "vpnClientNetworkRange": ["10.155.144.0/20"]
                vpn_type = vpnp.get("vpnType", "")
                vpn_gateway = vpnp.get("vpnGatewayAddress", "")
                client_addrs = vpnp.get("vpnClientAddresses", [])
                client_ranges = vpnp.get("vpnClientNetworkRange", [])

                # Build a concise string
                # e.g. "vpnType=cisco-anyconnect,gateway=165.214.12.240,client=10.155.159.173,ranges=10.155.144.0/20"
                info_str = (
                    f"vpnType={vpn_type},"
                    f"gateway={vpn_gateway},"
                    f"client={';'.join(client_addrs)},"
                    f"ranges={';'.join(client_ranges)}"
                )
                vpn_info_list.append(info_str)

            # Build final dict for this agent
            record = {
                "id": agent.get("id"),
                "name": agent.get("name"),
                "computerName": agent.get("computerName"),
                "osVersion": agent.get("osVersion"),
                "platform": agent.get("platform"),
                "lastSeen": agent.get("lastSeen"),
                "status": agent.get("status"),
                "deleted": agent.get("deleted"),
                "version": agent.get("version"),
                "createdAt": agent.get("createdAt"),
                "numberOfClients": agent.get("numberOfClients"),
                "locationName": location_name,
                "agentType": agent.get("agentType"),
                "licenseType": agent.get("licenseType"),
                # Our new fields
                "usernames": ",".join(sorted(username_set)),
                "localIpv4": ",".join(sorted(local_ipv4_set)),
                "gatewayIpv4": ",".join(sorted(gateway_set)),
                "hardwareTypes": ",".join(sorted(hardware_set)),
                "vpnInfo": "|".join(vpn_info_list),  # or some other delimiter
            }

            all_records.append(record)

        next_url = next_link
        page_count += 1

    return pd.DataFrame(all_records)


def is_private_ipv4(ip: str) -> bool:
    """
    Quick helper to check if 'ip' is in a private IPv4 range.
    This is a simplistic approach:
      - 10.x.x.x
      - 172.16.x.x - 172.31.x.x
      - 192.168.x.x
    For robust logic, consider 'ipaddress' module in Python.
    """
    if not ip:
        return False
    parts = ip.split(".")
    if len(parts) != 4:
        return False
    try:
        p0, p1, p2, p3 = map(int, parts)
    except ValueError:
        return False

    # Check well-known private ranges
    if p0 == 10:
        return True
    if p0 == 192 and p1 == 168:
        return True
    if p0 == 172 and 16 <= p1 <= 31:
        return True
    return False


def fetch_enterprise_tests(aid: str) -> pd.DataFrame:
    """
    Fetch the list of Enterprise Tests for a given Account ID.
    Endpoint: /tests?aid=XXXXXXX
    """
    url = f"{BASE_URL}/tests?aid={aid}"
    logger.info(f"Fetching Enterprise Tests for AID={aid} from '{url}'...")
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()

    data = resp.json().get("tests", [])
    logger.info(f"Received {len(data)} Enterprise Tests for AID={aid}.")
    records = []
    for test in data:
        records.append({
            "testID": test.get("testId"),
            "testName": test.get("testName"),
            "createdBy": test.get("createdBy"),
            "createdDate": test.get("createdDate"),
            "modifiedBy": test.get("modifiedBy"),
            "modifiedDate": test.get("modifiedDate"),
            "type": test.get("type"),
            "alertsEnabled": test.get("alertsEnabled"),
            "enabled": test.get("enabled"),
            "direction": test.get("direction"),
            "targetAgentID": test.get("targetAgentId")
        })
    return pd.DataFrame(records)


def fetch_scheduled_tests(aid: str) -> pd.DataFrame:
    """
    Fetch the scheduled endpoint tests for a given Account ID.
    Endpoint: /endpoint/tests/scheduled-tests?aid=XXXXXXX
    """
    url = f"{BASE_URL}/endpoint/tests/scheduled-tests?aid={aid}"
    logger.info(f"Fetching Scheduled Tests for AID={aid} from '{url}'...")
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()

    data = resp.json().get("tests", [])
    logger.info(f"Received {len(data)} Scheduled Tests for AID={aid}.")

    records = []
    for test in data:
        # Pull out agentSelectorConfig
        agent_selector_config = test.get("agentSelectorConfig", {})
        agent_selector_type = agent_selector_config.get("agentSelectorType", "")

        # If 'agentSelectorType' == "specific-agents"
        #   then there's a list of agent IDs under `agents`
        specific_agents = agent_selector_config.get("agents", [])

        # If 'agentSelectorType' == "agent-labels"
        #   then there's a list of label IDs under `endpointAgentLabels`
        endpoint_labels = agent_selector_config.get("endpointAgentLabels", [])

        # Everything else is the standard fields
        records.append({
            "testID": test.get("testId"),
            "testName": test.get("testName"),
            "server": test.get("server"),
            "createdDate": test.get("createdDate"),
            "type": test.get("type"),
            "isEnabled": test.get("isEnabled"),
            "agentSelectorType": agent_selector_type,
            # Join arrays into comma-separated strings for easy viewing in Excel
            "assignedAgentsRaw": ",".join(specific_agents) if specific_agents else "",
            "assignedLabelsRaw": ",".join(endpoint_labels) if endpoint_labels else "",
            # You can also store maxMachines if needed
            "maxMachines": agent_selector_config.get("maxMachines", None)
        })

    return pd.DataFrame(records)


def fetch_labels(aid: str) -> pd.DataFrame:
    """
    Fetch endpoint labels for a given Account ID, with filter details expanded,
    including agent-id, username, local-network, vpn-vendor, connection, etc.

    Endpoint: /endpoint/labels?aid={AID}&expand=filters
    """
    url = f"{BASE_URL}/endpoint/labels?aid={aid}&expand=filters"
    logger.info(f"Fetching Labels for AID={aid} from '{url}'...")
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()

    data = resp.json().get("labels", [])
    logger.info(f"Received {len(data)} Labels for AID={aid}.")

    records = []
    for label in data:
        label_id = label.get("id")
        name = label.get("name")
        color = label.get("color")
        match_type = label.get("matchType", "")

        # Track various filter keys. Now includes 'agent-id'.
        filter_info = {
            "agent-id": [],
            "username": [],
            "local-network": [],
            "vpn-vendor": [],
            "connection": []
        }

        # Parse the 'filters' array
        for fobj in label.get("filters", []):
            key = fobj.get("key", "").lower()
            vals = fobj.get("values", [])
            if key in filter_info:
                filter_info[key].extend(vals)

        # Create one record for this label
        row = {
            "id": label_id,
            "name": name,
            "color": color,
            "matchType": match_type,
            "agent_id_filter": ",".join(filter_info["agent-id"]),
            "local_network_filter": ",".join(filter_info["local-network"]),
            "vpn_vendor_filter": ",".join(filter_info["vpn-vendor"]),
            "connection_filter": ",".join(filter_info["connection"]),
            "username_filter": ",".join(filter_info["username"]),
        }
        records.append(row)

    df = pd.DataFrame(records)
    return df


def fetch_usage(aid: str):
    """
    Calls /usage?aid={AID}&expand=endpoint-agent&expand=test&expand=enterprise-agent
    Returns a dict of DataFrames:
        {
          "summary": usage_summary_df,
          "tests"  : usage_tests_df,
          "endpoint_agents": usage_endpoint_agents_df,
          "enterprise_agents": usage_enterprise_agents_df   # optional
        }
    If certain expansions are missing data, some DataFrames may be empty.
    """
    url = (
        f"{BASE_URL}/usage?aid={aid}"
        f"&expand=endpoint-agent&expand=test&expand=enterprise-agent"
    )
    logger.info(f"Fetching Usage info for AID={aid} from '{url}'...")
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()

    data = resp.json()
    usage_obj = data.get("usage", {})

    # 1) Usage Summary
    #    We store top-level usage and quota fields in a single-row DataFrame
    quota = usage_obj.get("quota", {})
    usage_summary = {
        "aid"                           : aid,
        "monthStart"                    : quota.get("monthStart"),
        "monthEnd"                      : quota.get("monthEnd"),
        "cloudUnitsIncluded"            : quota.get("cloudUnitsIncluded"),
        "deviceAgentsIncluded"          : quota.get("deviceAgentsIncluded"),
        "enterpriseAgentsIncluded"      : quota.get("enterpriseAgentsIncluded"),
        "endpointAgentsIncluded"        : quota.get("endpointAgentsIncluded"),
        "endpointAgentsEssentialsIncluded": quota.get("endpointAgentsEssentialsIncluded"),
        "cloudUnitsUsed"                : usage_obj.get("cloudUnitsUsed"),
        "cloudUnitsProjected"           : usage_obj.get("cloudUnitsProjected"),
        "cloudUnitsNextBillingPeriod"   : usage_obj.get("cloudUnitsNextBillingPeriod"),
        "enterpriseUnitsUsed"           : usage_obj.get("enterpriseUnitsUsed"),
        "enterpriseUnitsProjected"      : usage_obj.get("enterpriseUnitsProjected"),
        "enterpriseUnitsNextBillingPeriod" : usage_obj.get("enterpriseUnitsNextBillingPeriod"),
        "enterpriseAgentsUsed"          : usage_obj.get("enterpriseAgentsUsed"),
        "endpointAgentsUsed"            : usage_obj.get("endpointAgentsUsed"),
        "endpointAgentsEssentialsUsed"  : usage_obj.get("endpointAgentsEssentialsUsed"),
        "connectedDevicesUnitsUsed"     : usage_obj.get("connectedDevicesUnitsUsed"),
        "connectedDevicesUnitsProjected": usage_obj.get("connectedDevicesUnitsProjected"),
        "connectedDevicesUnitsNextBillingPeriod": usage_obj.get("connectedDevicesUnitsNextBillingPeriod"),
    }
    usage_summary_df = pd.DataFrame([usage_summary])  # single row

    # 2) Usage Tests
    tests_list = usage_obj.get("tests", [])
    usage_tests_records = []
    for t in tests_list:
        usage_tests_records.append({
            "aid"             : t.get("aid"),
            "testId"          : t.get("testId"),
            "accountGroupName": t.get("accountGroupName"),
            "testName"        : t.get("testName"),
            "testType"        : t.get("testType"),
            "cloudUnitsUsed"  : t.get("cloudUnitsUsed"),
            "cloudUnitsProjected": t.get("cloudUnitsProjected")
        })
    usage_tests_df = pd.DataFrame(usage_tests_records)

    # 3) Usage Endpoint Agents
    endpoint_agents_list = usage_obj.get("endpointAgents", [])
    usage_endpoint_agents_records = []
    for ea in endpoint_agents_list:
        usage_endpoint_agents_records.append({
            "aid"              : ea.get("aid"),
            "accountGroupName" : ea.get("accountGroupName"),
            "endpointAgentsUsed": ea.get("endpointAgentsUsed")
        })
    usage_endpoint_agents_df = pd.DataFrame(usage_endpoint_agents_records)

    # 4) Usage Enterprise Agents
    #    The sample JSON might have enterpriseAgents in the same "usage" object.
    #    If it's under a different key or combined, adapt accordingly.
    #    The sample snippet shows enterprise agents info was appended inside "endpointAgents" or something.
    #    If there's a separate key "enterpriseAgents", do something like below:
    enterprise_agents_list = usage_obj.get("enterpriseAgents", [])
    usage_enterprise_agents_records = []
    for ea in enterprise_agents_list:
        usage_enterprise_agents_records.append({
            "aid"                : ea.get("aid"),
            "accountGroupName"   : ea.get("accountGroupName"),
            "enterpriseAgentsUsed": ea.get("enterpriseAgentsUsed")
        })
    usage_enterprise_agents_df = pd.DataFrame(usage_enterprise_agents_records)

    return {
        "summary"           : usage_summary_df,
        "tests"             : usage_tests_df,
        "endpoint_agents"   : usage_endpoint_agents_df,
        "enterprise_agents" : usage_enterprise_agents_df
    }


def main():
    logger.info("Starting ThousandEyes data collection script...")

    # 1) Read account groups from local Excel file
    account_groups_df = get_account_ids("account_ids.xlsx")
    account_groups_tab = account_groups_df.copy()

    # We'll store final results in a dict of DataFrames (sheet_name -> DataFrame)
    sheets = {"Account Groups": account_groups_tab}

    # Prepare lists to aggregate data
    agents_records = []
    endpoint_agents_records = []
    enterprise_tests_records = []
    scheduled_tests_records = []
    labels_records = []

    # 2) For each account group, fetch data
    # (No usage call inside this loop anymore!)
    for idx, row in account_groups_df.iterrows():
        account_group_name = row["accountGroupName"]
        aid = str(row["aid"])
        logger.info(f"Processing accountGroupName='{account_group_name}', AID={aid}...")

        # ---- A) Enterprise Agents ----
        agents_df = fetch_agents(aid)
        if not agents_df.empty:
            agents_df.insert(0, "aid", aid)
            agents_df.insert(0, "accountGroupName", account_group_name)
            agents_records.append(agents_df)

        # ---- B) Endpoint Agents ----
        endpoint_agents_df = fetch_endpoint_agents(aid)
        if not endpoint_agents_df.empty:
            endpoint_agents_df.insert(0, "aid", aid)
            endpoint_agents_df.insert(0, "accountGroupName", account_group_name)
            endpoint_agents_records.append(endpoint_agents_df)

        # ---- C) Enterprise Tests ----
        enterprise_tests_df = fetch_enterprise_tests(aid)
        if not enterprise_tests_df.empty:
            enterprise_tests_df.insert(0, "aid", aid)
            enterprise_tests_df.insert(0, "accountGroupName", account_group_name)
            enterprise_tests_records.append(enterprise_tests_df)

        # ---- D) Scheduled Tests (Endpoint) ----
        scheduled_df = fetch_scheduled_tests(aid)
        if not scheduled_df.empty:
            scheduled_df.insert(0, "aid", aid)
            scheduled_df.insert(0, "accountGroupName", account_group_name)
            scheduled_tests_records.append(scheduled_df)

        # ---- E) Labels ----
        labels_df = fetch_labels(aid)
        if not labels_df.empty:
            labels_df.insert(0, "aid", aid)
            labels_df.insert(0, "accountGroupName", account_group_name)
            labels_records.append(labels_df)

    logger.info("Finished fetching data from all accounts. Consolidating data...")

    # 3) Concatenate or create empty dataframes for each category:

    # ---- A) Enterprise Agents (Tab: Agents) ----
    if agents_records:
        agents_final_df = pd.concat(agents_records, ignore_index=True)
    else:
        agents_final_df = pd.DataFrame(
            columns=[
                "accountGroupName", "aid", "OrgId", "agentId", "agentName",
                "agentType", "agentState", "lastSeen", "createdDate",
                "utilization", "location", "enabled", "hostname", "ipAddresses"
            ]
        )
    sheets["Agents"] = agents_final_df

    # ---- B) Endpoint Agents (Tab: Endpoint Agents) ----
    if endpoint_agents_records:
        endpoint_agents_final_df = pd.concat(endpoint_agents_records, ignore_index=True)
    else:
        endpoint_agents_final_df = pd.DataFrame(
            columns=[
                "accountGroupName", "aid", "id", "name", "computerName", "osVersion",
                "platform", "lastSeen", "status", "deleted", "version", "createdAt",
                "numberOfClients", "locationName", "agentType", "licenseType",
                "usernames", "localIpv4", "gatewayIpv4", "hardwareTypes", "vpnInfo"
            ]
        )

    desired_order = [
        "accountGroupName",
        "aid",
        "id",
        "name",
        "computerName",
        "osVersion",
        "platform",
        "lastSeen",
        "status",
        "deleted",
        "version",
        "createdAt",
        "numberOfClients",
        "locationName",
        "agentType",
        "licenseType",
        "usernames",
        "localIpv4",
        "gatewayIpv4",
        "hardwareTypes",
        "vpnInfo"
    ]
    endpoint_agents_final_df = endpoint_agents_final_df.reindex(columns=desired_order, fill_value="")
    sheets["Endpoint Agents"] = endpoint_agents_final_df

    # ---- C) Enterprise Tests (Tab: Enterprise Test) ----
    if enterprise_tests_records:
        enterprise_tests_final_df = pd.concat(enterprise_tests_records, ignore_index=True)
    else:
        enterprise_tests_final_df = pd.DataFrame(
            columns=[
                "accountGroupName", "aid", "testID", "testName", "createdBy",
                "createdDate", "modifiedBy", "modifiedDate", "type", "alertsEnabled",
                "enabled", "direction", "targetAgentID"
            ]
        )
    sheets["Enterprise Test"] = enterprise_tests_final_df

    # ---- D) Scheduled Test Endpoint Agent ----
    if scheduled_tests_records:
        scheduled_tests_final_df = pd.concat(scheduled_tests_records, ignore_index=True)
    else:
        scheduled_tests_final_df = pd.DataFrame(
            columns=[
                "accountGroupName", "aid", "testID", "testName", "server",
                "createdDate", "type", "isEnabled",
                "agentSelectorType",
                "assignedAgentsRaw",
                "assignedLabelsRaw"
            ]
        )
    sheets["Scheduled Test Endpoint Agent"] = scheduled_tests_final_df

    # ---- E) Labels ----
    if labels_records:
        labels_final_df = pd.concat(labels_records, ignore_index=True)
    else:
        labels_final_df = pd.DataFrame(
            columns=[
                "accountGroupName", "aid", "id", "name", "color", "matchType",
                "agent_id_filter", "username_filter", "local_network_filter",
                "vpn_vendor_filter", "connection_filter"
            ]
        )
    label_columns_desired = [
        "accountGroupName", "aid", "id", "name", "color", "matchType",
        "local_network_filter", "vpn_vendor_filter", "connection_filter",
        "username_filter", "agent_id_filter"
    ]
    labels_final_df = labels_final_df.reindex(columns=label_columns_desired, fill_value="")
    sheets["Labels"] = labels_final_df

    # ============= Correlation Steps =============
    logger.info("Building Label->Agent mapping and final Test->Agent assignments...")

    if not labels_final_df.empty and not endpoint_agents_final_df.empty:
        label_agents_map = build_label_agents_map(labels_final_df, endpoint_agents_final_df)
    else:
        label_agents_map = {}

    if not scheduled_tests_final_df.empty and not endpoint_agents_final_df.empty:
        test_agents_map = build_test_agents_map(
            scheduled_tests_final_df,
            label_agents_map,
            endpoint_agents_final_df
        )
        assignments_df = build_test_agent_assignment_df(
            scheduled_tests_final_df,
            test_agents_map,
            endpoint_agents_final_df
        )
        sheets["Test ↔ Agent Assignments"] = assignments_df
    else:
        assignments_df = pd.DataFrame(columns=["testID", "testName", "agentID", "agentName"])
        sheets["Test ↔ Agent Assignments"] = assignments_df

    # ============= SINGLE USAGE CALL =============
    # We'll pick the first row's AID from the account_groups_df
    if not account_groups_df.empty:
        top_level_aid = str(account_groups_df.iloc[0]["aid"])
        logger.info(f"Fetching usage data once using AID={top_level_aid}...")
        usage_data = fetch_usage(top_level_aid)

        # parse the usage_data dict -> DataFrames
        usage_summary_df = usage_data["summary"]
        usage_tests_df = usage_data["tests"]
        usage_endpoint_df = usage_data["endpoint_agents"]
        usage_ent_agents_df = usage_data["enterprise_agents"]

        # optional: Insert "accountGroupName" or "aid" columns if you want
        # usage_summary_df["aid"] = top_level_aid
        # usage_tests_df["aid"] = top_level_aid
        # usage_endpoint_df["aid"] = top_level_aid
        # usage_ent_agents_df["aid"] = top_level_aid

        # Place them in the sheets dict
        sheets["Usage Summary"] = usage_summary_df
        sheets["Usage Tests"] = usage_tests_df
        sheets["Usage Endpoint Agents"] = usage_endpoint_df
        sheets["Usage Enterprise Agents"] = usage_ent_agents_df
    else:
        # no account groups => no usage
        sheets["Usage Summary"] = pd.DataFrame()
        sheets["Usage Tests"] = pd.DataFrame()
        sheets["Usage Endpoint Agents"] = pd.DataFrame()
        sheets["Usage Enterprise Agents"] = pd.DataFrame()

    # 4) Write all sheets to a single Excel file
    timestamp_int = int(time.time())
    output_file = f"thousandeyes_data-{timestamp_int}.xlsx"
    logger.info(f"Writing all results to Excel file '{output_file}'...")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    logger.success(f"Data successfully written to '{output_file}'. Now applying color fills and number formats...")

    # 5) Color fills in "Labels"
    _apply_color_fills(output_file, sheet_name="Labels", color_col="color")

    # 6) Auto-format numeric columns in all sheets
    _auto_format_numbers(output_file)

    logger.success("All post-processing complete. Script finished.")


if __name__ == "__main__":
    main()
